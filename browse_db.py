# browse_db.py
import sqlite3
import json
import argparse
from datetime import datetime
from flask import Flask, render_template, request, jsonify, abort, send_from_directory, Response, send_file
from markupsafe import Markup 
import argparse
import tempfile
import os
import sys
import threading
import webbrowser
from collections import Counter
import rcssmin
import rjsmin
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill 
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.utils import secure_filename # Ensure this is imported
import gzip
import base64
import zstandard as zstd
import tarfile
import shutil

# Import globals, the classification and verification modules
import globals
import automate_classification
import verify_classification

# Define default year range - For this app:
DEFAULT_YEAR_FROM = 2011
DEFAULT_YEAR_TO = 2025
DEFAULT_MIN_PAGE_COUNT = 4

app = Flask(__name__)
DATABASE = None # Will be set from command line argument

def render_papers_table(hide_offtopic_param=None, year_from_param=None, year_to_param=None, min_page_count_param=None):
    """Fetches papers based on filters and renders the papers_table.html template. 
       Used for initial render from / and XHR updates."""
    # Determine hide_offtopic state
    hide_offtopic = True # Default
    if hide_offtopic_param is not None:
        hide_offtopic = hide_offtopic_param.lower() in ['1', 'true', 'yes', 'on']

    # Determine filter values, using defaults if not provided or invalid
    year_from_value = int(year_from_param) if year_from_param is not None else DEFAULT_YEAR_FROM
    year_to_value = int(year_to_param) if year_to_param is not None else DEFAULT_YEAR_TO
    min_page_count_value = int(min_page_count_param) if min_page_count_param is not None else DEFAULT_MIN_PAGE_COUNT

    # Fetch papers with ALL the filters applied
    papers = fetch_papers(
        hide_offtopic=hide_offtopic,
        year_from=year_from_value,
        year_to=year_to_value,
        min_page_count=min_page_count_value,
    )

    # Render the table template fragment, passing the search query value for the input field
    rendered_table = render_template(
        'papers_table.html',
        papers=papers,
        type_emojis=globals.TYPE_EMOJIS,
        default_type_emoji=globals.DEFAULT_TYPE_EMOJI,
        pdf_emojis=globals.PDF_EMOJIS, # Pass the PDF emojis dictionary
        hide_offtopic=hide_offtopic,
        # Pass the *string representations* of the values to the template for input fields
        year_from_value=str(year_from_value),
        year_to_value=str(year_to_value),
        min_page_count_value=str(min_page_count_value)
    )
    return rendered_table

# DB functions:
def get_db_connection():
    """Create a connection to the SQLite database and ensure FTS tables."""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row 
    return conn

def fetch_papers(hide_offtopic=True, year_from=None, year_to=None, min_page_count=None):
    """Fetch papers from the database, applying various optional filters."""
    conn = get_db_connection()
    base_query = "SELECT p.* FROM papers p"
    conditions = []
    params = []

    if hide_offtopic:
        conditions.append("(p.is_offtopic = 0 OR p.is_offtopic IS NULL)")
    if year_from is not None:
        try:
            year_from = int(year_from)
            conditions.append("p.year >= ?")
            params.append(year_from)
        except (ValueError, TypeError):
            pass
    if year_to is not None:
        try:
            year_to = int(year_to)
            conditions.append("p.year <= ?")
            params.append(year_to)
        except (ValueError, TypeError):
            pass
    if min_page_count is not None:
        try:
            min_page_count = int(min_page_count)
            conditions.append("(p.page_count IS NULL OR p.page_count = '' OR p.page_count >= ?)")
            params.append(min_page_count)
        except (ValueError, TypeError):
            pass

    # --- Build Final Query ---
    # Start with base query
    query_parts = [base_query]

    # Add WHERE clause if conditions exist
    if conditions:
        query_parts.append("WHERE " + " AND ".join(conditions))

    # Add ORDER BY clause for user comments first
    # This sorts rows where user_trace is NOT NULL and NOT empty string first
    query_parts.append("ORDER BY (p.user_trace IS NULL OR p.user_trace = '') ASC")

    # Combine all parts
    query = " ".join(query_parts)

    try:
        papers = conn.execute(query, params).fetchall()
    except sqlite3.Error as e:
        conn.close()
        print(f"Database error during fetch_papers: {e}")
        raise # Re-raise to be caught by the calling function (e.g., render_papers_table)
    finally:
        conn.close()

    # --- Process Results (Same as before) ---
    paper_list = []
    for paper in papers:
        paper_dict = dict(paper)
        try:
            paper_dict['features'] = json.loads(paper_dict['features'])
        except (json.JSONDecodeError, TypeError):
            paper_dict['features'] = {}
        try:
            paper_dict['technique'] = json.loads(paper_dict['technique'])
        except (json.JSONDecodeError, TypeError):
            paper_dict['technique'] = {}

        paper_dict['pdf_filename'] = paper_dict.get('pdf_filename')     # Could be None or a string
        paper_dict['pdf_state'] = paper_dict.get('pdf_state', 'none')   # Default state if not present
        paper_dict['changed_formatted'] = format_changed_timestamp(paper_dict.get('changed'))
        # paper_dict['authors_truncated'] = truncate_authors(paper_dict.get('authors', ''))
        paper_list.append(paper_dict)
    
    return paper_list

def update_paper_custom_fields(paper_id, data, changed_by="user"):
    """Update the custom classification fields for a paper and audit fields.
       Handles partial updates based on keys present in `data`."""

    conn = get_db_connection()
    cursor = conn.cursor()
    
    changed_timestamp = datetime.utcnow().isoformat() + 'Z'

    update_fields = []
    update_values = []

    # Handle Main Boolean Fields (Partial Update)
    main_bool_fields = ['is_survey', 'is_offtopic', 'is_through_hole', 'is_smt', 'is_x_ray']
    for field in main_bool_fields:
        if field in data:
            value = data[field]
            if isinstance(value, str):
                if value.lower() in ('true', '1', 'on'):
                    update_fields.append(f"{field} = ?")
                    update_values.append(1)
                elif value.lower() in ('false', '0'):
                    update_fields.append(f"{field} = ?")
                    update_values.append(0)
                else: # 'unknown', '', None, etc.
                    update_fields.append(f"{field} = ?")
                    update_values.append(None)
            elif value is True:
                update_fields.append(f"{field} = ?")
                update_values.append(1)
            elif value is False:
                update_fields.append(f"{field} = ?")
                update_values.append(0)
            else: # value is None or numeric
                update_fields.append(f"{field} = ?")
                update_values.append(int(bool(value)) if value is not None else None)

    # Handle Research Area (Partial Update)
    if 'research_area' in data:
        update_fields.append("research_area = ?")
        update_values.append(data['research_area'])


    page_count_value_for_pages_update = None # Variable to hold the value for potential 'pages' update
    if 'page_count' in data:
        page_count_value = data['page_count']
        if page_count_value is not None:
            try:
                page_count_value = int(page_count_value)
                # Store the integer value for potential 'pages' update
                page_count_value_for_pages_update = page_count_value 
            except (ValueError, TypeError):
                page_count_value = None
                page_count_value_for_pages_update = None # Reset if invalid
        else:
            page_count_value_for_pages_update = None # Reset if None
        
        update_fields.append("page_count = ?")
        update_values.append(page_count_value)

        cursor.execute("SELECT pages FROM papers WHERE id = ?", (paper_id,))
        row = cursor.fetchone()
        if row:
            current_pages_value = row['pages']
            # Check if 'pages' is effectively empty/blank/null
            # This checks for None, empty string, or string with only whitespace
            if current_pages_value is None or (isinstance(current_pages_value, str) and current_pages_value.strip() == ""):
                # If 'pages' is blank and we have a valid page_count to set
                if page_count_value_for_pages_update is not None:
                    # Add the update for the 'pages' column
                    update_fields.append("pages = ?")
                    # Convert the integer page count back to string for the 'pages' TEXT column
                    update_values.append(str(page_count_value_for_pages_update)) 
        # --- END NEW LOGIC ---

    # Handle Verified By (Partial Update)
    if 'verified_by' in data:
        verified_by_value = data['verified_by']
        # Ensure value is either 'user' or None. Others (like model names) are treated as None.
        # This enforces that the UI can only set 'user' or clear it.
        if verified_by_value != 'user':
            verified_by_value = None
        update_fields.append("verified_by = ?")
        update_values.append(verified_by_value)
        

    # Handle Features (Partial Update)
    # Fetch current features JSON from DB to merge changes
    cursor.execute("SELECT features FROM papers WHERE id = ?", (paper_id,))
    row = cursor.fetchone()
    if row:
        try:
            current_features = json.loads(row['features']) if row['features'] else {}
        except (json.JSONDecodeError, TypeError):
            current_features = {}
    else:
        current_features = {}

    # Check for feature fields in the incoming data
    feature_updates = {}
    for key in list(data.keys()): # Iterate over a copy to safely modify data
        if key.startswith('features_'):
            feature_key = key.split('features_')[1]
            value = data.pop(key) # Remove from main data dict
            if feature_key == 'other':
                feature_updates[feature_key] = value # Text field
            else:
                # Handle radio button group for 3-state (true/false/unknown)
                if isinstance(value, str):
                    if value == 'true':
                        feature_updates[feature_key] = True
                    elif value == 'false':
                        feature_updates[feature_key] = False
                    else: # 'unknown' or anything else
                        feature_updates[feature_key] = None
                elif isinstance(value, bool):
                    feature_updates[feature_key] = value
                else: # None or numeric
                    feature_updates[feature_key] = bool(value) if value is not None else None

    if feature_updates:
        current_features.update(feature_updates)
        update_fields.append("features = ?")
        update_values.append(json.dumps(current_features))

    # Handle Techniques (Partial Update)
    # Fetch current technique JSON from DB to merge changes
    cursor.execute("SELECT technique FROM papers WHERE id = ?", (paper_id,))
    row = cursor.fetchone()
    if row:
        try:
            current_technique = json.loads(row['technique']) if row['technique'] else {}
        except (json.JSONDecodeError, TypeError):
            current_technique = {}
    else:
        current_technique = {}

    # Check for technique fields in the (potentially modified) data
    technique_updates = {}
    for key in list(data.keys()):
        if key.startswith('technique_'):
            technique_key = key.split('technique_')[1]
            value = data.pop(key) # Remove from main data dict
            if technique_key == 'model':
                technique_updates[technique_key] = value # Text field
            elif technique_key == 'available_dataset':
                 # Handle radio button group for 3-state (true/false/unknown)
                if isinstance(value, str):
                    if value == 'true':
                        technique_updates[technique_key] = True
                    elif value == 'false':
                        technique_updates[technique_key] = False
                    else: # 'unknown' or anything else
                        technique_updates[technique_key] = None
                elif isinstance(value, bool):
                    technique_updates[technique_key] = value
                else: # None or numeric
                    technique_updates[technique_key] = bool(value) if value is not None else None
            else: # classic_computer_vision_based, machine_learning_based, hybrid
                 # Handle radio button group for 3-state (true/false/unknown)
                if isinstance(value, str):
                    if value == 'true':
                        technique_updates[technique_key] = True
                    elif value == 'false':
                        technique_updates[technique_key] = False
                    else: # 'unknown' or anything else
                        technique_updates[technique_key] = None
                elif isinstance(value, bool):
                    technique_updates[technique_key] = value
                else: # None or numeric
                    technique_updates[technique_key] = bool(value) if value is not None else None

    # Merge updates into current technique
    if technique_updates:
        current_technique.update(technique_updates)
        update_fields.append("technique = ?")
        update_values.append(json.dumps(current_technique))

    # Always update audit fields for any change
    update_fields.append("changed = ?")
    update_values.append(changed_timestamp)
    update_fields.append("changed_by = ?")
    update_values.append(changed_by)

    # Any remaining keys in 'data' are assumed to be direct column names
    for key, value in data.items(): # Iterate over the potentially modified data dict
        # Skip keys already handled or special keys
        if key in ['id', 'changed', 'changed_by', 'verified_by'] or key in main_bool_fields or key.startswith(('features_', 'technique_')):
            continue
        # Treat remaining keys as direct column updates
        update_fields.append(f"{key} = ?")
        update_values.append(value)

    if update_fields:
        update_query = f"UPDATE papers SET {', '.join(update_fields)} WHERE id = ?"
        update_values.append(paper_id)
        # --- Debug prints ---
        # print(f"DEBUG: Updating paper {paper_id}")
        # print(f"DEBUG: SQL Query: {update_query}")
        # print(f"DEBUG: Values: {update_values}")
        # --- End Debug prints ---
        cursor.execute(update_query, update_values)
        conn.commit()
        rows_affected = cursor.rowcount
    else:
        rows_affected = 0 # No fields to update
    conn.close()

    if rows_affected > 0:
        conn = get_db_connection()
        updated_paper = conn.execute("SELECT * FROM papers WHERE id = ?", (paper_id,)).fetchone()
        conn.close()
        
        if updated_paper:
            updated_dict = dict(updated_paper)
            try:
                updated_dict['features'] = json.loads(updated_dict['features'])
            except (json.JSONDecodeError, TypeError):
                updated_dict['features'] = {}
            try:
                updated_dict['technique'] = json.loads(updated_dict['technique'])
            except (json.JSONDecodeError, TypeError):
                updated_dict['technique'] = {}
            
            updated_dict['changed_formatted'] = format_changed_timestamp(updated_dict.get('changed'))
            

            return_data = {
                'status': 'success',
                'changed': updated_dict.get('changed'),
                'changed_formatted': updated_dict['changed_formatted'],
                'changed_by': updated_dict.get('changed_by'),
                # Include updated fields for frontend refresh
                'research_area': updated_dict.get('research_area'),
                'page_count': updated_dict.get('page_count'),
                'is_survey': updated_dict.get('is_survey'),
                'is_offtopic': updated_dict.get('is_offtopic'),
                'is_through_hole': updated_dict.get('is_through_hole'),
                'is_smt': updated_dict.get('is_smt'),
                'is_x_ray': updated_dict.get('is_x_ray'),
                'relevance': updated_dict.get('relevance'),
                'features': updated_dict['features'], # Parsed dict
                'technique': updated_dict['technique'], # Parsed dict
                'user_trace': updated_dict.get('user_trace')
            }
            return return_data
        else:
            return {'status': 'error', 'message': 'Paper not found after update.'}
    else:
        return {'status': 'error', 'message': 'No rows updated. Paper ID might not exist or no changes were made.'}

def fetch_updated_paper_data(paper_id):
    """Fetches the full paper data after classification/verification for client-side update."""
    conn = get_db_connection()
    try:
        updated_paper = conn.execute("SELECT * FROM papers WHERE id = ?", (paper_id,)).fetchone()
        if updated_paper:
            updated_dict = dict(updated_paper)
            try:
                updated_dict['features'] = json.loads(updated_dict['features'])
            except (json.JSONDecodeError, TypeError):
                updated_dict['features'] = {}
            try:
                updated_dict['technique'] = json.loads(updated_dict['technique'])
            except (json.JSONDecodeError, TypeError):
                updated_dict['technique'] = {}
            updated_dict['changed_formatted'] = format_changed_timestamp(updated_dict.get('changed'))
            
            # Prepare data for frontend refresh (matching update_paper_custom_fields structure)
            return_data = {
                'status': 'success',
                'changed': updated_dict.get('changed'),
                'changed_formatted': updated_dict['changed_formatted'],
                'changed_by': updated_dict.get('changed_by'),
                'verified_by': updated_dict.get('verified_by'),
                # Include updated fields for frontend refresh
                'research_area': updated_dict.get('research_area'),
                'page_count': updated_dict.get('page_count'),
                'is_survey': updated_dict.get('is_survey'),
                'is_offtopic': updated_dict.get('is_offtopic'),
                'is_through_hole': updated_dict.get('is_through_hole'),
                'is_smt': updated_dict.get('is_smt'),
                'is_x_ray': updated_dict.get('is_x_ray'),
                'relevance': updated_dict.get('relevance'),
                'verified': updated_dict.get('verified'),
                'estimated_score': updated_dict.get('estimated_score'),
                'features': updated_dict['features'], # Parsed dict
                'technique': updated_dict['technique'], # Parsed dict
                'reasoning_trace': updated_dict.get('reasoning_trace'), # Include traces
                'verifier_trace': updated_dict.get('verifier_trace'),
                'user_trace': updated_dict.get('user_trace')
            }
            return return_data
        else:
            return {'status': 'error', 'message': 'Paper not found after update.'}
    finally:
        conn.close()

# --- DB Helper Functions ---

def format_changed_timestamp(changed_str):
    """Format the ISO timestamp string to dd/mm/yy hh:mm:ss"""
    if not changed_str:
        return ""
    try:
        dt = datetime.fromisoformat(changed_str.replace('Z', '+00:00'))
        return dt.strftime("%d/%m/%y %H:%M:%S")
    except ValueError:
        # If parsing fails, return the original string or a placeholder
        return changed_str

# used for HTML and XLSX exports: 
def get_default_filter_values(hide_offtopic_param, year_from_param, year_to_param, min_page_count_param):
    """Extracts and validates filter parameters, returning default values if invalid/missing."""
    hide_offtopic = True 
    if hide_offtopic_param is not None:
        hide_offtopic = hide_offtopic_param.lower() in ['1', 'true', 'yes', 'on']

    try:
        year_from_value = int(year_from_param) if year_from_param is not None else DEFAULT_YEAR_FROM
    except ValueError:
        year_from_value = DEFAULT_YEAR_FROM

    try:
        year_to_value = int(year_to_param) if year_to_param is not None else DEFAULT_YEAR_TO
    except ValueError:
        year_to_value = DEFAULT_YEAR_TO

    try:
        min_page_count_value = int(min_page_count_param) if min_page_count_param is not None else DEFAULT_MIN_PAGE_COUNT
    except ValueError:
        min_page_count_value = DEFAULT_MIN_PAGE_COUNT

    return hide_offtopic, year_from_value, year_to_value, min_page_count_value

# --- Core Export Generation Functions ---

def generate_html_export_content(papers, hide_offtopic, year_from_value, year_to_value, min_page_count_value, is_lite_export=False):
    """Generates the full HTML content string for the static export."""
    # Strip fat text for lite export:
    if is_lite_export:
        for paper in papers:
            paper['abstract'] = '' # Blank Abstract
            paper['reasoning_trace'] = '' # Blank Classifier Trace
            paper['verifier_trace'] = '' # Blank Verifier Trace

    fonts_css_content = ""
    style_css_content = ""
    chart_js_content = ""
    chart_js_datalabels_content = ""
    d3_js_content = ""
    d3_cloud_js_content = ""
    ghpages_js_content = ""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    static_dir = os.path.join(script_dir, 'static')
    try:
        with open(os.path.join(static_dir, 'libs/chart.min.js'), 'r', encoding='utf-8') as f:
            chart_js_content = f.read()
        with open(os.path.join(static_dir, 'libs/chartjs-plugin-datalabels.min.js'), 'r', encoding='utf-8') as f:
            chart_js_datalabels_content = f.read()
        with open(os.path.join(static_dir, 'libs/d3.min.js'), 'r', encoding='utf-8') as f:
            d3_js_content = f.read()
        with open(os.path.join(static_dir, 'libs/d3-cloud.min.js'), 'r', encoding='utf-8') as f:
            d3_cloud_js_content = f.read()

        with open(os.path.join(static_dir, 'fonts.css'), 'r', encoding='utf-8') as f:
            fonts_css_content = f.read()
        with open(os.path.join(static_dir, 'style.css'), 'r', encoding='utf-8') as f:
            style_css_content = f.read()
        with open(os.path.join(static_dir, 'ghpages.js'), 'r', encoding='utf-8') as f:
            ghpages_js_content = f.read()
        with open(os.path.join(static_dir, 'stats.js'), 'r', encoding='utf-8') as f:
            stats_js_content = f.read()
        with open(os.path.join(static_dir, 'filtering.js'), 'r', encoding='utf-8') as f:
            filtering_js_content = f.read()
    except FileNotFoundError as e:
        print(f"Warning: Static file not found during HTML export generation: {e}")
        # Handle missing files gracefully if possible, or raise an error
        raise

    fonts_css_content = rcssmin.cssmin(fonts_css_content)
    style_css_content = rcssmin.cssmin(style_css_content)
    
    chart_js_content = rjsmin.jsmin(chart_js_content)
    chart_js_datalabels_content = rjsmin.jsmin(chart_js_datalabels_content)
    d3_js_content = rjsmin.jsmin(d3_js_content)
    d3_cloud_js_content = rjsmin.jsmin(d3_cloud_js_content)

    stats_js_content = rjsmin.jsmin(stats_js_content)
    filtering_js_content = rjsmin.jsmin(filtering_js_content)
    ghpages_js_content = rjsmin.jsmin(ghpages_js_content)

    # --- Render the static export template ---
    papers_table_static_export = render_template(
        'papers_table_static_export.html',
        papers=papers, # Pass the potentially modified papers list
        type_emojis=globals.TYPE_EMOJIS,
        pdf_emojis=globals.PDF_EMOJIS, # Pass the PDF emojis dictionary
        default_type_emoji=globals.DEFAULT_TYPE_EMOJI,
        hide_offtopic=hide_offtopic,
        year_from_value=str(year_from_value),
        year_to_value=str(year_to_value),
        min_page_count_value=str(min_page_count_value),
    )
    full_html_content = render_template(
        'index_static_export.html',
        papers_table_static_export=papers_table_static_export,
        hide_offtopic=hide_offtopic,
        year_from_value=year_from_value, # Pass raw values if needed by template logic
        year_to_value=year_to_value,
        min_page_count_value=min_page_count_value,
        # --- Pass potentially minified static content ---
        fonts_css_content=Markup(fonts_css_content), # Markup was already applied if needed, or content is minified
        style_css_content=Markup(style_css_content),
        
        chart_js_content=Markup(chart_js_content),
        chart_js_datalabels_content=Markup(chart_js_datalabels_content),
        d3_js_content=Markup(d3_js_content),
        d3_cloud_js_content=Markup(d3_cloud_js_content),

        filtering_js_content=Markup(filtering_js_content),
        stats_js_content=Markup(stats_js_content),
        ghpages_js_content=Markup(ghpages_js_content)
    )

    # --- Compress the full HTML content ---
    html_bytes = full_html_content.encode('utf-8')  # 1. Encode the HTML string to bytes (UTF-8)
    compressed_bytes = gzip.compress(html_bytes)    # 2. Compress the bytes
    compressed_base64 = base64.b64encode(compressed_bytes).decode('ascii')  # 3. Encode the compressed bytes to Base64 for embedding in JS

    pako_js_content = ""
    try:
        with open(os.path.join(static_dir, 'libs/pako.min.js'), 'r', encoding='utf-8') as f:
            pako_js_content = f.read()
    except FileNotFoundError as e:
        print(f"Warning: pako.min.js not found during HTML export generation: {e}")
        # Handle missing pako.js gracefully or raise an error
        raise

    # --- Render the LOADER template, passing the compressed data ---
    loader_html_content = render_template(
        'loader.html',
        compressed_html_data=compressed_base64,
        pako_js_content=Markup(pako_js_content)
    )
    return loader_html_content

def generate_xlsx_export_content(papers):
    """Generates the Excel file content as bytes."""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "PCB Inspection Papers"

    # --- Define Headers (Updated Order - Corrected Boolean Features) ---
    headers = [
        "Type", "Title", "Year", "Journal/Conf name", "Pages count",
        # Classification Summary
        "Off-topic", "Relevance", "Survey", "THT", "SMT", "X-Ray",
        # Features Summary (Updated Order - Corrected Boolean Features)
        "Tracks", "Holes / Vias", "Bare PCB Other", # Boolean (e.g., bare_pcb_other)
        "Solder Insufficient", "Solder Excess", "Solder Void", "Solder Crack", "Solder Other", # Boolean (e.g., solder_other)
        "Missing Comp", "Wrong Comp", "Orientation", "Comp Other", # Boolean (e.g., component_other)
        "Cosmetic", "Other State", # Boolean for state (based on 'other' text content)
        "Other Defects Text", # Text for content (the 'other' field)
        # Techniques Summary (Updated Order)
        "Classic CV", "ML", "CNN Classifier", "CNN Detector",
        "R-CNN Detector", "Transformers", "Other DL", "Hybrid", "Datasets", "Model name",
        # Metadata
        "Last Changed", "Changed By", "Verified", "Accr. Score", "Verified By",
        "User Comment State", "User Comments" # Boolean for state, Text for content
    ]

    # --- Write Headers ---
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # --- Write Data Rows ---
    for row_num, paper in enumerate(papers, 2): # Start from row 2
        # --- Helper function for consistent Excel value conversion ---
        def format_excel_value(val):
            """
            Converts Python/DB values to Excel-friendly values:
            - True/1   -> TRUE (Excel boolean)
            - False/0  -> FALSE (Excel boolean)
            - None/''/etc. -> "" (Empty string for blank Excel cell)
            - Other    -> str(val) (Text)
            """
            if val is True or (isinstance(val, (int, float)) and val == 1):
                return True # Excel TRUE
            elif val is False or (isinstance(val, (int, float)) and val == 0):
                return False # Excel FALSE
            elif val is None or val == "":
                 return "" # Explicitly empty cell for NULL/empty
            else:
                # Handle potential string representations of booleans from inconsistent DB
                if isinstance(val, str):
                    lower_val = val.lower()
                    if lower_val in ('true', '1'):
                        return True
                    elif lower_val in ('false', '0'):
                        return False
                # Default: Convert to string for text fields
                return str(val)

        # Extract and format data
        features = paper.get('features', {})
        technique = paper.get('technique', {})

        # --- Format the 'Last Changed' date ---
        changed_timestamp_str = paper.get('changed', '')
        formatted_changed_date = ""
        if changed_timestamp_str:
            try:
                # Parse the ISO format timestamp
                dt = datetime.fromisoformat(changed_timestamp_str.replace('Z', '+00:00'))
                # Format as 'YYYY-MM-DD HH:MM:SS' for Excel compatibility
                formatted_changed_date = dt.strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                # If parsing fails, keep the original string or leave blank
                formatted_changed_date = changed_timestamp_str # Or ""

        row_data = [
            paper.get('type', ''),                    # Type (text)
            paper.get('title', ''),                   # Title (text)
            paper.get('year'),                        # Year (integer)
            paper.get('journal', ''),                 # Journal/Conf name (text)
            paper.get('page_count'),                  # Pages count (integer)
            # --- Classification Summary ---
            format_excel_value(paper.get('is_offtopic')), # Off-topic (boolean/null)
            paper.get('relevance'),                   # Relevance (integer)
            format_excel_value(paper.get('is_survey')), # Survey (boolean/null)
            format_excel_value(paper.get('is_through_hole')), # THT (boolean/null)
            format_excel_value(paper.get('is_smt')),    # SMT (boolean/null)
            format_excel_value(paper.get('is_x_ray')),  # X-Ray (boolean/null)
            # --- Features Summary (Updated Order - Corrected Boolean Features) ---
            format_excel_value(features.get('tracks')), # Tracks (boolean/null)
            format_excel_value(features.get('holes')),  # Holes / Vias (boolean/null)
            format_excel_value(features.get('bare_pcb_other')), # Bare PCB Other (boolean/null) - ADDED
            format_excel_value(features.get('solder_insufficient')), # Solder Insufficient (boolean/null)
            format_excel_value(features.get('solder_excess')), # Solder Excess (boolean/null)
            format_excel_value(features.get('solder_void')), # Solder Void (boolean/null)
            format_excel_value(features.get('solder_crack')), # Solder Crack (boolean/null)
            format_excel_value(features.get('solder_other')), # Solder Other (boolean/null) - ADDED
            format_excel_value(features.get('missing_component')), # Missing Comp (boolean/null)
            format_excel_value(features.get('wrong_component')), # Wrong Comp (boolean/null)
            format_excel_value(features.get('orientation')), # Orientation (boolean/null)
            format_excel_value(features.get('component_other')), # Comp Other (boolean/null) - ADDED
            format_excel_value(features.get('cosmetic')), # Cosmetic (boolean/null)
            # Other State (boolean based on 'other' text content) - CORRECTED COMMENT
            format_excel_value(features.get('other') is not None and str(features.get('other', '')).strip() != ""),
            features.get('other', ''),               # Other Defects Text (text) - This one shows the text
            # --- Techniques Summary (Updated Order) ---
            format_excel_value(technique.get('classic_cv_based')), # Classic CV (boolean/null)
            format_excel_value(technique.get('ml_traditional')), # ML (boolean/null)
            format_excel_value(technique.get('dl_cnn_classifier')), # CNN Classifier (boolean/null)
            format_excel_value(technique.get('dl_cnn_detector')), # CNN Detector (boolean/null)
            format_excel_value(technique.get('dl_rcnn_detector')), # R-CNN Detector (boolean/null)
            format_excel_value(technique.get('dl_transformer')), # Transformers (boolean/null)
            format_excel_value(technique.get('dl_other')), # Other DL (boolean/null)
            format_excel_value(technique.get('hybrid')), # Hybrid (boolean/null)
            format_excel_value(technique.get('available_dataset')), # Datasets (boolean/null)
            technique.get('model', ''),              # Model name (text)
            # --- Metadata ---
            formatted_changed_date,                 # Last Changed (formatted date string)
            paper.get('changed_by', ''),            # Changed By (text)
            format_excel_value(paper.get('verified')), # Verified (boolean/null)
            paper.get('estimated_score'),           # Accr. Score (integer)
            paper.get('verified_by', ''),           # Verified By (text)
            # User comments state (boolean based on 'user_trace' text content) - CORRECTED COMMENT
            format_excel_value(paper.get('user_trace') is not None and str(paper.get('user_trace', '')).strip() != ""),
            paper.get('user_trace', '')             # User comments contents (text) - This one shows the text
        ]

        # Write the row data to Excel
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Optional: Auto-adjust column widths (basic attempt)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter # Get the column name
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        # Cap the width to prevent extremely wide columns
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

    # Optional: Format the data as a table (requires openpyxl >= 2.5)
    try:
        if len(papers) > 0:
            # Adjust the column reference to 'AQ' (assuming 37 columns now: A through AQ)
            # Headers are row 1, data starts row 2, so last row is len(papers) + 1
            tab = Table(displayName="PCBPapersTable", ref=f"A1:AQ{len(papers) + 1}")
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)
    except Exception as e:
        print(f"Warning: Could not create Excel table: {e}")

    # --- NEW: Apply Conditional Formatting for Boolean Cells ---
    # Define fills for TRUE and FALSE
    true_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid") # Light Green
    false_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Light Red
    # Updated boolean column indices based on corrected new order (1-based indexing)
    boolean_columns = [
        # Classification Summary
        6, 8, 9, 10, 11,
        # Features Summary (Boolean Features)
        12, 13, 14, # Tracks, Holes, Bare PCB Other
        15, 16, 17, 18, 19, # Solder Insufficient, Excess, Void, Crack, Solder Other
        20, 21, 22, 23, # Missing Comp, Wrong Comp, Orientation, Comp Other
        24, 25, 27, # Cosmetic, Other State, User Comment State
        # Techniques Summary
        28, 29, 30, 31, 32, 33, 34, 35, 36,
        # Metadata
        39 # Verified (column 39)
    ]

    # Iterate through rows and specified boolean columns to apply formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for col_idx in boolean_columns:
            # Adjust for 0-based indexing in the row list
            cell = row[col_idx - 1] # col_idx is 1-based, list index is 0-based
            if cell.value is True:
                cell.fill = true_fill
            elif cell.value is False:
                cell.fill = false_fill
            # If cell.value is None or "", it remains unformatted (blank cell)

    # --- Save Workbook to BytesIO object ---
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_filename(base_name, year_from, year_to, min_page_count, hide_offtopic, extra_suffix=""):
    """Generates a filename based on filters."""
    filename_parts = [base_name]
    if year_from == year_to:
        filename_parts.append(str(year_from))
    else:
        filename_parts.append(f"{year_from}-{year_to}")
    if min_page_count > 0:
        filename_parts.append(f"min{min_page_count}pg")
    if hide_offtopic:
        filename_parts.append("noOfftopic")
    if extra_suffix:
        filename_parts.append(extra_suffix)
    return "_".join(filename_parts)


# --- Jinja2-like filters ---
def render_status(value):
    """Render status value as emoji/symbol"""
    if value == 1 or value == "true" or value is True:
        return '‚úîÔ∏è' # Checkmark for True
    elif value == 0  or value == "false" or value is False:
        return '‚ùå' # Cross for False
    else: # None or unknown
        return '‚ùî' # Question mark for Unknown/Null

def render_verified_by(value):
    """
    Render verified_by value as emoji.
    Accepts the raw database value.
    Returns HTML string with emoji and tooltip if needed.
    """
    if value == 'user':
        return f'<span title="User">üë§</span>' # Human emoji
    elif value is None or value == '':
        return f'<span title="Unverified">‚ùî</span>'
    else:
        # For any other string, value is a model name, show computer emoji with tooltip
        # Escape the model name for HTML attribute safety
        escaped_model_name = str(value).replace('"', '&quot;').replace("'", "&#39;")
        return f'<span title="{escaped_model_name}">üñ•Ô∏è</span>'

def render_changed_by(value):
    """
    Render changed_by value as emoji.
    Accepts the raw database value.
    Returns HTML string with emoji and tooltip if needed.
    """
    if value == 'user':
        return f'<span title="User">üë§</span>' # Human emoji
    elif value is None or value == '':
        return f'<span title="Unknown">‚ùî</span>' # Question mark for null/empty
    else:
        # For any other string, value is a model name, show computer emoji with tooltip
        escaped_model_name = str(value).replace('"', '&quot;').replace("'", "&#39;")
        return f'<span title="{escaped_model_name}">üñ•Ô∏è</span>'

# --- BibTeX Generation Function ---
def generate_bibtex_string(paper):
    """
    Generates a raw BibTeX entry string from a paper dictionary.
    Excludes the abstract field.
    Handles different entry types and common fields.
    """
    if not paper or not paper.get('id'):
        return "Error: Paper ID missing."

    entry_type = paper.get('type', 'misc').lower() # Default to 'misc' if type is missing
    bibtex_key = paper['id'] # Use the database ID as the BibTeX key

    # Map common database fields to BibTeX fields
    field_mapping = {
        'title': paper.get('title'),
        'author': paper.get('authors'), # Assuming authors are stored as a semicolon-separated string, might need reformatting for BibTeX
        'year': paper.get('year'),
        'journal': paper.get('journal'),
        'volume': paper.get('volume'),
        'number': paper.get('number'), # Add number if it exists in DB (mapped from issue)
        'pages': paper.get('pages'), # Already normalized, might need further cleaning for BibTeX range format
        'doi': paper.get('doi'),
        'issn': paper.get('issn'),
        'month': paper.get('month'),
        # Add other fields as needed, e.g., 'booktitle' for 'inproceedings'
    }

    # Handle author formatting (semicolon-separated to ' and ' separated, potentially)
    # This is a simple conversion, might need refinement based on your exact storage format
    if field_mapping['author']:
        # Split by semicolon and join with ' and '
        authors_list = [author.strip() for author in field_mapping['author'].split(';')]
        field_mapping['author'] = ' and '.join(authors_list)

    # Handle pages formatting for BibTeX range (e.g., "123 - 125" -> "123--125")
    if field_mapping['pages']:
        # Simple replacement, might need more robust parsing if formats vary significantly
        # Remove spaces around hyphens/dashes and replace with double dash
        import re
        pages_str = field_mapping['pages']
        # Replace spaces around various dash-like characters with double hyphen
        field_mapping['pages'] = re.sub(r'\s*[-‚Äì‚Äî]\s*', '--', pages_str)

    # Determine fields relevant to the entry type and build the entry
    # This is a basic example, you might want to expand this mapping
    type_required_fields = {
        'article': ['title', 'author', 'journal', 'year'],
        'inproceedings': ['title', 'author', 'booktitle', 'year'], # Note: booktitle, not journal
        'book': ['title', 'author', 'publisher', 'year'],
        'inbook': ['title', 'author', 'chapter', 'publisher', 'year'],
        'incollection': ['title', 'author', 'booktitle', 'publisher', 'year'],
        'techreport': ['title', 'author', 'institution', 'year'],
        'phdthesis': ['title', 'author', 'school', 'year'],
        'mastersthesis': ['title', 'author', 'school', 'year'],
        'manual': ['title'],
        'misc': ['title'], # Default, might add author, howpublished, note
        'conference': ['title', 'author', 'booktitle', 'year'], # Treat like inproceedings
    }

    relevant_fields = type_required_fields.get(entry_type, ['title', 'author', 'year']) # Default fields

    # Start building the BibTeX string
    lines = [f"@{entry_type}{{{bibtex_key},"]
    for field in relevant_fields:
        value = field_mapping.get(field)
        if value is not None and str(value).strip() != "": # Only add non-empty fields
            # Basic escaping for common BibTeX characters if needed (e.g., #, {, }, $)
            # bibtex_value = str(value).replace('{', '\\{').replace('}', '\\}') # Example
            bibtex_value = str(value)
            # Ensure proper quoting, using double quotes as standard
            lines.append(f"  {field} = {{{bibtex_value}}},") # Use braces for safety

    # Add other potentially relevant fields that aren't strictly "required"
    other_fields = ['volume', 'number', 'pages', 'doi', 'issn', 'month']
    for field in other_fields:
        value = field_mapping.get(field)
        if value is not None and str(value).strip() != "" and field not in relevant_fields:
             bibtex_value = str(value)
             lines.append(f"  {field} = {{{bibtex_value}}},") # Use braces

    lines.append("}") # Close the entry
    return "\n".join(lines)

# --- Register the new filter ---
@app.template_filter('bibtex')
def bibtex_filter(paper):
    """Jinja2 filter to generate a BibTeX string for a paper dictionary."""
    return generate_bibtex_string(paper)

@app.template_filter('render_changed_by')
def render_changed_by_filter(value):
    # Use Markup to tell Jinja2 that the output is safe HTML
    return Markup(render_changed_by(value))

@app.template_filter('render_status')
def render_status_filter(value):
    return render_status(value)

@app.template_filter('render_verified_by')
def render_verified_by_filter(value):
    # Use Markup to tell Jinja2 that the output is safe HTML
    return Markup(render_verified_by(value)) 






#Routes: 
@app.route('/', methods=['GET'])
def index():
    """Main page to display the table."""
    # Get initial filter parameters from the request (or they will default inside render_papers_table)
    hide_offtopic_param = request.args.get('hide_offtopic')
    year_from_param = request.args.get('year_from')
    year_to_param = request.args.get('year_to')
    min_page_count_param = request.args.get('min_page_count')
        
    # Get the total number of papers in the database.
    conn = get_db_connection()
    total_paper_count = conn.execute("SELECT COUNT(*) FROM papers").fetchone()[0]
    conn.close()

    papers_table_content = render_papers_table(
        hide_offtopic_param=hide_offtopic_param,
        year_from_param=year_from_param,
        year_to_param=year_to_param,
        min_page_count_param=min_page_count_param,
    )
    # Pass the rendered table content and filter values to the main index template
    # Determine values to display in the input fields (use defaults if URL params were missing/invalid)
    try:
        year_from_input_value = str(int(year_from_param)) if year_from_param is not None else str(DEFAULT_YEAR_FROM)
    except ValueError:
        year_from_input_value = str(DEFAULT_YEAR_FROM)
    try:
        year_to_input_value = str(int(year_to_param)) if year_to_param is not None else str(DEFAULT_YEAR_TO)
    except ValueError:
        year_to_input_value = str(DEFAULT_YEAR_TO)
    try:
        min_page_count_input_value = str(int(min_page_count_param)) if min_page_count_param is not None else str(DEFAULT_MIN_PAGE_COUNT)
    except ValueError:
        min_page_count_input_value = str(DEFAULT_MIN_PAGE_COUNT)
    hide_offtopic_checkbox_checked = hide_offtopic_param is None or hide_offtopic_param.lower() in ['1', 'true', 'yes', 'on']

    return render_template(
        'index.html',
        papers_table_content=papers_table_content,
        hide_offtopic=hide_offtopic_checkbox_checked,
        year_from_value=year_from_input_value,
        year_to_value=year_to_input_value,
        min_page_count_value=min_page_count_input_value,
        total_paper_count=total_paper_count
    )

#Backup/restore
@app.route('/backup', methods=['GET'])
def backup_database():
    """Creates a backup of the database and related files."""
    try:
        # Create backup filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"{timestamp}.par√ßa.zst"

        # Create temporary directory for exports
        with tempfile.TemporaryDirectory() as temp_dir:
            # Generate HTML export (full, not lite)
            papers = fetch_papers(hide_offtopic=True, year_from=0, year_to=9999, min_page_count=0)
            html_content = generate_html_export_content(papers, True, 0, 9999, 0, is_lite_export=False)
            html_path = os.path.join(temp_dir, 'export.html')
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)

            # Generate XLSX export
            xlsx_content = generate_xlsx_export_content(papers)
            xlsx_path = os.path.join(temp_dir, 'export.xlsx')
            with open(xlsx_path, 'wb') as f:
                f.write(xlsx_content)

            # Create in-memory buffer for the backup
            buffer = io.BytesIO()
            
            # Create a Zstandard compressor
            cctx = zstd.ZstdCompressor(level=1)  # Fastest compression level
            
            # Compress the tar directly to the buffer
            with tarfile.open(fileobj=buffer, mode='w') as tar:
                # Add database file
                tar.add(DATABASE, arcname='data/new.sqlite')
                
                # Add PDF storage directory
                if os.path.exists(globals.PDF_STORAGE_DIR):
                    tar.add(globals.PDF_STORAGE_DIR, arcname='data/pdf')
                
                # Add annotated PDF storage directory
                if os.path.exists(globals.ANNOTATED_PDF_STORAGE_DIR):
                    tar.add(globals.ANNOTATED_PDF_STORAGE_DIR, arcname='data/pdf_annotated')
                
                # Add export files
                tar.add(html_path, arcname='export.html')
                tar.add(xlsx_path, arcname='export.xlsx')
            
            # Get the uncompressed tar data
            tar_data = buffer.getvalue()
            
            # Now compress the tar data with zstd
            compressed_data = cctx.compress(tar_data)
            
            # Create a new buffer with the compressed data
            compressed_buffer = io.BytesIO(compressed_data)
            compressed_buffer.seek(0)
            
            # Create a response with the in-memory backup
            response = send_file(
                compressed_buffer,
                as_attachment=True,
                download_name=backup_filename,
                mimetype='application/zstd'
            )
            
            # Ensure the filename is set correctly in Content-Disposition
            response.headers['Content-Disposition'] = f'attachment; filename="{backup_filename}"'
            
            return response
    except Exception as e:
        print(f"Backup error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    
@app.route('/restore', methods=['POST'])
def restore_database():
    """Restores database and related files from a backup."""
    try:
        if 'backup_file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No backup file provided'}), 400

        file = request.files['backup_file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400

        if not file.filename.endswith('.par√ßa.zst'):
            return jsonify({'status': 'error', 'message': 'Invalid backup file format. Expected .par√ßa.zst'}), 400

        # Create temporary directory for extraction
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save uploaded file temporarily
            temp_backup_path = os.path.join(temp_dir, 'backup.par√ßa.zst')
            file.save(temp_backup_path)

            # Decompress and extract
            dctx = zstd.ZstdDecompressor()
            with open(temp_backup_path, 'rb') as compressed_file:
                with dctx.stream_reader(compressed_file) as decomp_stream:
                    with tarfile.open(fileobj=decomp_stream, mode='r|') as tar:
                        tar.extractall(path=temp_dir)

            # Paths in the extracted archive
            extracted_db_path = os.path.join(temp_dir, 'data', 'new.sqlite')
            extracted_pdf_dir = os.path.join(temp_dir, 'data', 'pdf')
            extracted_annotated_pdf_dir = os.path.join(temp_dir, 'data', 'pdf_annotated')

            # Verify required files exist
            if not os.path.exists(extracted_db_path):
                return jsonify({'status': 'error', 'message': 'Backup does not contain required database file'}), 500

            # Backup current data before restoring (single file name, overwrites previous)
            backup_current = "backup_before_restore.par√ßa.zst"
            backup_current_path = os.path.join(os.getcwd(), backup_current)
            cctx = zstd.ZstdCompressor(level=1)
            with cctx.stream_writer(open(backup_current_path, 'wb')) as compressor:
                with tarfile.open(fileobj=compressor, mode='w|') as tar:
                    if os.path.exists(DATABASE):
                        tar.add(DATABASE, arcname='data/new.sqlite')
                    if os.path.exists(globals.PDF_STORAGE_DIR):
                        tar.add(globals.PDF_STORAGE_DIR, arcname='data/pdf')
                    if os.path.exists(globals.ANNOTATED_PDF_STORAGE_DIR):
                        tar.add(globals.ANNOTATED_PDF_STORAGE_DIR, arcname='data/pdf_annotated')

            # Perform restoration
            # 1. Replace database
            shutil.move(extracted_db_path, DATABASE)
            
            # 2. Replace PDF directories - only if they exist in the backup
            if os.path.exists(extracted_pdf_dir):
                # Create parent directory if it doesn't exist
                os.makedirs(os.path.dirname(globals.PDF_STORAGE_DIR), exist_ok=True)
                # Remove existing directory if it exists
                if os.path.exists(globals.PDF_STORAGE_DIR):
                    shutil.rmtree(globals.PDF_STORAGE_DIR)
                # Move the extracted directory
                shutil.move(extracted_pdf_dir, globals.PDF_STORAGE_DIR)
            else:
                # Create empty PDF directory if not in backup
                os.makedirs(globals.PDF_STORAGE_DIR, exist_ok=True)
                
            if os.path.exists(extracted_annotated_pdf_dir):
                # Create parent directory if it doesn't exist
                os.makedirs(os.path.dirname(globals.ANNOTATED_PDF_STORAGE_DIR), exist_ok=True)
                # Remove existing directory if it exists
                if os.path.exists(globals.ANNOTATED_PDF_STORAGE_DIR):
                    shutil.rmtree(globals.ANNOTATED_PDF_STORAGE_DIR)
                # Move the extracted directory
                shutil.move(extracted_annotated_pdf_dir, globals.ANNOTATED_PDF_STORAGE_DIR)
            else:
                # Create empty annotated PDF directory if not in backup
                os.makedirs(globals.ANNOTATED_PDF_STORAGE_DIR, exist_ok=True)

        return jsonify({
            'status': 'success',
            'message': f'Restored successfully from backup. Previous data backed up as {backup_current}'
        })
    except Exception as e:
        print(f"Restore error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# PDF storage/annotation routes
@app.route('/upload_pdf/<paper_id>', methods=['POST']) # Removed int: converter
def upload_pdf(paper_id):
    """Handles PDF file upload for a specific paper."""
    print(f"Received upload request for paper ID: {paper_id}") # Debug print
    if 'pdf_file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part in request'}), 400

    file = request.files['pdf_file']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'No file selected'}), 400

    if file and file.filename.lower().endswith('.pdf'):
        # original_filename = secure_filename(file.filename)
        unique_filename = f"{paper_id}.pdf"
        filepath = os.path.join(globals.PDF_STORAGE_DIR, unique_filename)

        try:
            file.save(filepath)
            # Update the database with the new filename and initial state 'PDF'
            # Use the string paper_id for the database query
            update_data = {'pdf_filename': unique_filename, 'pdf_state': 'PDF'}
            result = update_paper_custom_fields(paper_id, update_data, changed_by="user") # Pass string ID

            if result['status'] == 'success':
                # Fetch updated paper data including new PDF info
                # Pass the string ID here too
                updated_paper_data = fetch_updated_paper_data(paper_id)
                if updated_paper_data['status'] == 'success':
                    # Add the PDF specific data to the response
                    updated_paper_data['pdf_filename'] = unique_filename
                    updated_paper_data['pdf_state'] = 'PDF'
                    return jsonify(updated_paper_data)
                else:
                     print(f"Failed to fetch updated data for {paper_id} after upload.")
                     return jsonify({'status': 'error', 'message': 'Failed to fetch updated paper data after upload'}), 500
            else:
                print(f"DB update failed for {paper_id} after saving file.")
                # Rollback: delete the file if DB update failed
                if os.path.exists(filepath):
                    os.remove(filepath)
                return jsonify({'status': 'error', 'message': 'Failed to update database after saving file'}), 500

        except Exception as e:
            print(f"Error saving uploaded PDF for paper {paper_id}: {e}")
            # Attempt to remove the file if saving failed partway
            if os.path.exists(filepath):
                try:
                    os.remove(filepath)
                except OSError:
                    pass # Ignore error if file removal also fails
            return jsonify({'status': 'error', 'message': 'Failed to save file'}), 500
    else:
        print(f"File type not allowed for {paper_id}: {file.filename}")
        return jsonify({'status': 'error', 'message': 'File type not allowed, only PDFs are accepted'}), 400

@app.route('/serve_pdf/<paper_id>')
def serve_pdf(paper_id):
    """
    Serves the correct PDF file (annotated or original) for the PDF.js viewer/annotator
    based on the paper_id. Also updates the pdf_state in the database
    based on the actual existence of the annotated files.
    """
    conn = get_db_connection()
    paper = conn.execute("SELECT pdf_filename, pdf_state FROM papers WHERE id = ?", (paper_id,)).fetchone()
    
    if not paper or not paper['pdf_filename']:
        conn.close()
        print(f"No PDF filename found in DB for paper_id: {paper_id}")
        abort(404)

    filename = paper['pdf_filename']
    current_db_state = paper['pdf_state']
    
    # NEW LOGIC: Check for annotated and original files
    annotated_path = os.path.join(globals.ANNOTATED_PDF_STORAGE_DIR, filename)
    original_path = os.path.join(globals.PDF_STORAGE_DIR, filename)

    print(f"Serving PDF for paper_id {paper_id} (filename: {filename})") # Debug print
    print(f"Looking for annotated: {annotated_path}") # Debug print
    print(f"Looking for original: {original_path}") # Debug print

    new_state = None
    file_to_serve = None

    if os.path.exists(annotated_path):
        print(f"Found annotated file: {filename}")
        new_state = 'annotated'
        file_to_serve = annotated_path
    elif os.path.exists(original_path):
        print(f"Found original file: {filename}")
        new_state = 'PDF' # Reset to 'PDF' if annotated file is missing but original exists
        file_to_serve = original_path
    else:
        print(f"No PDF file found for paper_id: {paper_id} (filename: {filename})")
        # File doesn't exist at all, set state to 'none'
        new_state = 'none'
        # Update the database record
        update_query = "UPDATE papers SET pdf_state = ? WHERE id = ?"
        conn.execute(update_query, (new_state, paper_id))
        conn.commit()
        conn.close()
        abort(404) # Still abort 404 as no file to serve

    # Check if the state in the DB needs updating
    if current_db_state != new_state:
        print(f"Updating pdf_state for {paper_id} from '{current_db_state}' to '{new_state}'")
        update_query = "UPDATE papers SET pdf_state = ? WHERE id = ?"
        conn.execute(update_query, (new_state, paper_id))
        conn.commit()
    else:
        print(f"pdf_state for {paper_id} is already correct ('{new_state}')")

    conn.close()

    # Serve the determined file
    # Use os.path.basename to get just the filename from the full path for send_from_directory
    return send_from_directory(os.path.dirname(file_to_serve), os.path.basename(file_to_serve), as_attachment=False)

@app.route('/upload_annotated_pdf/<paper_id>', methods=['POST'])
def upload_annotated_pdf(paper_id):
    """
    API call for annotator autosaving feature:
    Receives an annotated PDF file associated with a paper_id,
    saves it to the annotated storage directory, and updates the pdf_state.
    """
    if 'pdf_file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part in request'}), 400

    file = request.files['pdf_file']
    if file.filename == '' or not file.filename.lower().endswith('.pdf'):
        return jsonify({'status': 'error', 'message': 'Invalid or missing file'}), 400

    # Look up filename from paper_id to save it correctly.
    conn = get_db_connection()
    paper = conn.execute("SELECT pdf_filename FROM papers WHERE id = ?", (paper_id,)).fetchone()
    conn.close()

    if not paper or not paper['pdf_filename']:
        return jsonify({'status': 'error', 'message': f'Paper ID {paper_id} not found in DB'}), 404
    
    # Use the filename from the DB. Sanitize for security.
    filename = secure_filename(paper['pdf_filename'])
    filepath = os.path.join(globals.ANNOTATED_PDF_STORAGE_DIR, filename) # [2]

    try:
        file.save(filepath)
        
        # Update the database to set the pdf_state to 'annotated'
        update_data = {'pdf_state': 'annotated'}
        result = update_paper_custom_fields(paper_id, update_data, changed_by="user") # [10, 11]
        
        if result.get('status') != 'success':
             # If DB update fails, you might want to remove the saved file to avoid inconsistency.
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'status': 'error', 'message': 'Failed to update paper state in DB'}), 500

        print(f"Saved annotated PDF for paper {paper_id} as {filename}")
        return jsonify({'status': 'success', 'message': f'File {filename} updated successfully.'})
    except Exception as e:
        print(f"Error saving annotated PDF for paper {paper_id}: {e}")
        return jsonify({'status': 'error', 'message': 'Failed to save file on server.'}), 500

# Export routes
@app.route('/static_export', methods=['GET'])
def static_export():
    """Generate and serve a downloadable HTML snapshot based on current filters."""
    # --- Get filter parameters from the request (URL query params) ---
    hide_offtopic_param = request.args.get('hide_offtopic')
    year_from_param = request.args.get('year_from')
    year_to_param = request.args.get('year_to')
    min_page_count_param = request.args.get('min_page_count')
    # hidden params (usable, but not implemented in the Web client GUI):
    lite_param = request.args.get('lite', default='0')
    download_param = request.args.get('download', default='1')

    # --- Determine filter values ---
    hide_offtopic, year_from_value, year_to_value, min_page_count_value= get_default_filter_values(
        hide_offtopic_param, year_from_param, year_to_param, min_page_count_param
    )

    # --- Fetch papers based on these filters ---
    papers = fetch_papers(
        hide_offtopic=hide_offtopic,
        year_from=year_from_value,
        year_to=year_to_value,
        min_page_count=min_page_count_value,
    )

    is_lite_export = lite_param.lower() in ['1', 'true', 'yes']

    # --- Generate the content using the core function ---
    full_html_content = generate_html_export_content(
        papers, hide_offtopic, year_from_value, year_to_value, min_page_count_value, is_lite_export
    )

    # --- Create a filename based on filters ---
    extra_suffix = "lite" if is_lite_export else ""
    filename = generate_filename("ResearchPar√ßa", year_from_value, year_to_value, min_page_count_value, hide_offtopic, extra_suffix) + ".html"

    # --- Prepare Response Headers ---
    response_headers = {"Content-Type": "text/html"}
    if download_param == '0':       # If download=0, set Content-Disposition to 'inline' to display in browser
        response_headers["Content-Disposition"] = f"inline; filename={filename}"
        print(f"Serving static export inline: {filename}") # Optional: Log action
    else:                           # Default behavior: prompt for download
        response_headers["Content-Disposition"] = f"attachment; filename={filename}"
        print(f"Sending static export as attachment: {filename}") # Optional: Log action

    return Response(
        full_html_content,
        mimetype="text/html",
        headers=response_headers
    )

@app.route('/xlsx_export', methods=['GET'])
def export_excel():
    """Generate and serve a downloadable Excel (.xlsx) file based on current filters."""
    # --- Get filter parameters from the request (URL query params) ---
    hide_offtopic_param = request.args.get('hide_offtopic')
    year_from_param = request.args.get('year_from')
    year_to_param = request.args.get('year_to')
    min_page_count_param = request.args.get('min_page_count')

    # --- Determine filter values ---
    hide_offtopic, year_from_value, year_to_value, min_page_count_value = get_default_filter_values(
        hide_offtopic_param, year_from_param, year_to_param, min_page_count_param
    )

    # --- Fetch papers based on these filters ---
    papers = fetch_papers(
        hide_offtopic=hide_offtopic,
        year_from=year_from_value,
        year_to=year_to_value,
        min_page_count=min_page_count_value,
    )

    # --- Generate the content using the core function ---
    excel_bytes = generate_xlsx_export_content(papers)

    # --- Create a filename based on filters ---
    filename = generate_filename("ResearchPar√ßa", year_from_value, year_to_value, min_page_count_value, hide_offtopic) + ".xlsx"

    # --- Return as a downloadable attachment ---
    return Response(
        excel_bytes, # Get the bytes from the core function
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Table generation routes
@app.route('/get_detail_row', methods=['GET'])
def get_detail_row():
    """Endpoint to fetch and render the detail row content for a specific paper."""
    paper_id = request.args.get('paper_id')
    if not paper_id:
        return jsonify({'status': 'error', 'message': 'Paper ID is required'}), 400

    try:
        conn = get_db_connection()
        paper = conn.execute("SELECT * FROM papers WHERE id = ?", (paper_id,)).fetchone()
        conn.close()

        if paper:
            # Process the paper data like in fetch_papers for consistency
            paper_dict = dict(paper)
            try:
                paper_dict['features'] = json.loads(paper_dict['features'])
            except (json.JSONDecodeError, TypeError):
                paper_dict['features'] = {}
            try:
                paper_dict['technique'] = json.loads(paper_dict['technique'])
            except (json.JSONDecodeError, TypeError):
                paper_dict['technique'] = {}
            
            # Render the detail row template fragment for this specific paper
            detail_html = render_template('detail_row.html', paper=paper_dict)
            return jsonify({'status': 'success', 'html': detail_html})
        else:
            return jsonify({'status': 'error', 'message': 'Paper not found'}), 404
    except Exception as e:
        print(f"Error fetching detail row for paper {paper_id}: {e}")
        return jsonify({'status': 'error', 'message': 'Failed to fetch detail row'}), 500

@app.route('/load_table', methods=['GET'])
def load_table():
    """Endpoint to fetch and render the table content based on current filters."""
    # Get filter parameters from the request
    hide_offtopic_param = request.args.get('hide_offtopic')
    year_from_param = request.args.get('year_from')
    year_to_param = request.args.get('year_to')
    min_page_count_param = request.args.get('min_page_count')

    # Use the updated helper function to render the table, passing the search query
    table_html = render_papers_table(
        hide_offtopic_param=hide_offtopic_param,
        year_from_param=year_from_param,
        year_to_param=year_to_param,
        min_page_count_param=min_page_count_param,
    )
    return table_html

# Data import/update routes (data writing):
@app.route('/update_paper', methods=['POST'])
def update_paper():
    """Endpoint to handle AJAX updates (partial or full)."""
    data = request.get_json()
    paper_id = data.get('id')
    if not paper_id:
        return jsonify({'status': 'error', 'message': 'Paper ID is required'}), 400

    try:
        # Use 'user' as the identifier for changes made via this interface
        result = update_paper_custom_fields(paper_id, data, changed_by="user")
        # The result dict already contains status and other data
        return jsonify(result)
    except Exception as e:
        print(f"Error updating paper {paper_id}: {e}") # Log error
        return jsonify({'status': 'error', 'message': 'Failed to update database'}), 500

# LLM inference routes (used for both single-paper and batch actions):
@app.route('/classify', methods=['POST'])
def classify_paper():
    """Endpoint to handle classification requests (single or batch)."""
    data = request.get_json()
    mode = data.get('mode', 'id') # Default to 'id' for single paper
    paper_id = data.get('paper_id')
    
    # Determine DB file (use command-line arg or global default)
    db_file = DATABASE 

    def run_classification_task():
        """Background task to run classification."""
        try:
            print(f"Starting classification task: mode={mode}, paper_id={paper_id}")
            # Call the appropriate function from automate_classification
            # Pass db_file explicitly or rely on its internal defaults/globals
            automate_classification.run_classification(
                mode=mode,
                paper_id=paper_id,
                db_file=db_file
                # grammar_file=..., prompt_template=..., server_url=... # Use defaults or override if needed
            )
            print(f"Classification task completed: mode={mode}, paper_id={paper_id}")
        except Exception as e:
            print(f"Error during background classification (mode={mode}, paper_id={paper_id}): {e}")
            # Consider logging this error more formally

    if mode in ['all', 'remaining']:
        # Run batch classification in a background thread to avoid blocking
        thread = threading.Thread(target=run_classification_task)
        thread.daemon = True # Dies with main process
        thread.start()
        # Return immediately
        return jsonify({'status': 'started', 'message': f'Batch classification ({mode}) initiated.'})
    elif mode == 'id' and paper_id:
        try:
            # Run single paper classification synchronously
            # The function updates the DB directly
            automate_classification.run_classification(
                mode=mode,
                paper_id=paper_id,
                db_file=db_file
            )
            # Fetch the updated data from the database
            updated_data = fetch_updated_paper_data(paper_id)
            if updated_data['status'] == 'success':
                return jsonify(updated_data)
            else:
                return jsonify(updated_data), 404 # Or 500 if it's a server error fetching data
        except Exception as e:
            print(f"Error classifying paper {paper_id}: {e}")
            return jsonify({'status': 'error', 'message': f'Classification failed: {str(e)}'}), 500
    else:
        return jsonify({'status': 'error', 'message': 'Invalid mode or missing paper_id for single classification.'}), 400

@app.route('/verify', methods=['POST'])
def verify_paper():
    """Endpoint to handle verification requests (single or batch)."""
    data = request.get_json()
    mode = data.get('mode', 'id') # Default to 'id' for single paper
    paper_id = data.get('paper_id')
    
    # Determine DB file (use command-line arg or global default)
    db_file = DATABASE

    def run_verification_task():
        """Background task to run verification."""
        try:
            print(f"Starting verification task: mode={mode}, paper_id={paper_id}")
            # Call the appropriate function from verify_classification
            verify_classification.run_verification(
                mode=mode,
                paper_id=paper_id,
                db_file=db_file
            )
            print(f"Verification task completed: mode={mode}, paper_id={paper_id}")
        except Exception as e:
            print(f"Error during background verification (mode={mode}, paper_id={paper_id}): {e}")
            # Consider logging this error more formally

    if mode in ['all', 'remaining']:
        # Run batch verification in a background thread to avoid blocking
        thread = threading.Thread(target=run_verification_task)
        thread.daemon = True
        thread.start()
        # Return immediately
        return jsonify({'status': 'started', 'message': f'Batch verification ({mode}) initiated.'})
    elif mode == 'id' and paper_id:
        try:
            # Run single paper verification synchronously
            verify_classification.run_verification(
                mode=mode,
                paper_id=paper_id,
                db_file=db_file
            )
            # Fetch the updated data from the database
            updated_data = fetch_updated_paper_data(paper_id)
            if updated_data['status'] == 'success':
                return jsonify(updated_data)
            else:
                return jsonify(updated_data), 404 # Or 500
        except Exception as e:
            print(f"Error verifying paper {paper_id}: {e}")
            return jsonify({'status': 'error', 'message': f'Verification failed: {str(e)}'}), 500
    else:
        return jsonify({'status': 'error', 'message': 'Invalid mode or missing paper_id for single verification.'}), 400

@app.route('/upload_bibtex', methods=['POST'])
def upload_bibtex():
    """Endpoint to handle BibTeX file upload and import."""
    global DATABASE # Assuming DATABASE is defined globally as before

    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file part'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'No selected file'}), 400

    if file and file.filename.lower().endswith('.bib'):
        try:
            # Save the uploaded file to a temporary location
            with tempfile.NamedTemporaryFile(delete=False, suffix='.bib') as tmp_bib_file:
                file.save(tmp_bib_file.name)
                tmp_bib_path = tmp_bib_file.name

            # Use the existing import_bibtex logic
            # Import here to avoid potential circular imports if placed at the top
            import import_bibtex 

            # Call the import function with the temporary file and the global DB path
            import_bibtex.import_bibtex(tmp_bib_path, DATABASE)

            # Clean up the temporary file
            os.unlink(tmp_bib_path)

            return jsonify({'status': 'success', 'message': 'BibTeX file imported successfully.'})

        except Exception as e:
            # Ensure cleanup even if import fails
            if 'tmp_bib_path' in locals():
                try:
                    os.unlink(tmp_bib_path)
                except OSError:
                    pass # Ignore errors during cleanup
            print(f"Error importing BibTeX: {e}")
            return jsonify({'status': 'error', 'message': f'Import failed: {str(e)}'}), 500
    else:
        return jsonify({'status': 'error', 'message': 'Invalid file type. Please upload a .bib file.'}), 400



if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Browse and edit PCB inspection papers database.')
    parser.add_argument('db_file', nargs='?', help='SQLite database file path (optional)')
    args = parser.parse_args()
    
    if args.db_file:
        DATABASE = args.db_file
        print(f"Attempting to use database file from command line argument: {DATABASE}")
    elif hasattr(globals, 'DATABASE_FILE') and globals.DATABASE_FILE:
        DATABASE = globals.DATABASE_FILE
        print(f"Attempting to use database file from globals.DATABASE_FILE: {DATABASE}")

    # 3. If DATABASE is still None (neither arg nor globals provided), or if the specified file doesn't exist,
    #    fall back to 'fallback.sqlite' by copying it to globals.DATABASE_FILE location
    fallback_needed = False
    if DATABASE is None:
        fallback_needed = True
        print("Info: No database file specified via argument or globals.DATABASE_FILE.")
    elif not os.path.exists(DATABASE):
        fallback_needed = True
        print(f"Warning: Specified database file not found: {DATABASE}")

    if fallback_needed:
        # Check if fallback.sqlite exists in the script's directory
        script_dir = os.path.dirname(os.path.abspath(__file__))
        fallback_path = os.path.join(script_dir, 'fallback.sqlite')
        
        if not os.path.exists(fallback_path):
            print(f"Error: Fallback database file not found: {fallback_path}")
            print("Please ensure 'fallback.sqlite' exists in the script's directory.")
            sys.exit(1)
        
        target_database = globals.DATABASE_FILE
        print(f"Copying fallback database from {fallback_path} to {target_database}")
        
        # Copy the fallback database to the target location
        import shutil
        shutil.copy2(fallback_path, target_database)
        
        DATABASE = target_database
        print(f"Using database file: {DATABASE}")

    # Check if the final determined database file exists
    if not os.path.exists(DATABASE):
        print(f"Error: Final database file not found: {DATABASE}")
        print("Please provide a valid database file via command line argument, set globals.DATABASE_FILE correctly, or ensure 'fallback.sqlite' exists in the script's directory.")
        sys.exit(1) # Exit if even the fallback doesn't exist

    # Verify the database has the required tables
    try:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='papers'")
        if not cursor.fetchone():
            print(f"Error: Database '{DATABASE}' does not contain required 'papers' table")
            sys.exit(1)
        conn.close()
    except sqlite3.Error as e:
        print(f"Error verifying database: {e}")
        sys.exit(1)

    print(f"Starting server, database: {DATABASE}")

    # --- Open browser only once ---
    # The standard Flask/Werkzeug reloader runs the script twice:
    # 1. Once in the parent process (to manage the reloader)
    # 2. Once in the child process (the actual server, where WERKZEUG_RUN_MAIN is set)
    # We only want to open the browser in the child process.
    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        # Function to open the browser after a delay
        def open_browser():
            import time
            time.sleep(2)  # Wait for the server to start
            webbrowser.open("http://127.0.0.1:5000")

        # Start the browser opener in a separate thread
        threading.Thread(target=open_browser, daemon=True).start()
        print(" * Visit http://127.0.0.1:5000 to view the table.")
    
    # Ensure the templates and static folders exist
    if not os.path.exists('templates'):
        os.makedirs('templates')
    if not os.path.exists('static'):
        os.makedirs('static')
    app.run(host='0.0.0.0', port=5000, debug=True)